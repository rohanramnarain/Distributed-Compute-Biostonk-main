"""Convert clinical-trial embedding workbooks into one compressed dataset."""

import argparse
import logging
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

REQUIRED_COLUMNS = {"nct_id", "sentiment"}
VALID_SENTIMENTS = {"0.0", "1.0"}
REQUIRED_NPZ_KEYS = ["nct_id", "label_names", "source_workbook"]


def validate_npz_schema(data_dict: dict[str, Any]) -> tuple[bool, str]:
    """Validates that loaded/processed dataset elements conform to the expected schema keys.

    Returns:
        Tuple[bool, str]: Success boolean and descriptive status message.
    """
    missing_keys = [key for key in REQUIRED_NPZ_KEYS if key not in data_dict]
    if missing_keys:
        msg = f"Schema Validation Error: Missing required keys {missing_keys}"
        logger.warning(msg)
        return False, msg

    if len(data_dict["nct_id"]) == 0:
        msg = "Schema Validation Error: Generated dataset is empty."
        logger.warning(msg)
        return False, msg

    return True, "Schema validation passed."


def read_workbook(path: Path) -> tuple[list[str], np.ndarray, list[str], int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    header = next(rows)
    columns = {str(name): index for index, name in enumerate(header) if name is not None}
    missing = REQUIRED_COLUMNS - columns.keys()
    if missing:
        raise ValueError(f"{path.name} is missing columns: {', '.join(sorted(missing))}")

    embedding_columns = sorted(
        (name for name in columns if name.startswith("emb_")),
        key=lambda name: int(name.removeprefix("emb_")),
    )
    if not embedding_columns:
        raise ValueError(f"{path.name} has no embedding columns")

    trial_ids: list[str] = []
    embeddings: list[list[float]] = []
    sentiments: list[str] = []
    skipped_rows = 0
    for row in rows:
        required_indexes = [columns["nct_id"], columns["sentiment"]]
        required_indexes.extend(columns[column] for column in embedding_columns)
        if len(row) <= max(required_indexes):
            skipped_rows += 1
            continue
        nct_id = row[columns["nct_id"]]
        sentiment = row[columns["sentiment"]]
        if nct_id is None or str(sentiment) not in VALID_SENTIMENTS:
            skipped_rows += 1
            continue
        embedding = [row[columns[column]] for column in embedding_columns]
        if any(value is None for value in embedding):
            skipped_rows += 1
            continue
        trial_ids.append(str(nct_id))
        embeddings.append([float(value) for value in embedding])
        sentiments.append(str(sentiment))

    return trial_ids, np.asarray(embeddings, dtype=np.float32), sentiments, skipped_rows


def main(input_directory: Path, output_path: Path) -> None:
    workbooks = sorted(input_directory.glob("*.xlsx"))
    if not workbooks:
        raise FileNotFoundError(f"No .xlsx files found in {input_directory}")

    all_ids: list[str] = []
    all_embeddings: list[np.ndarray] = []
    all_sentiments: list[str] = []
    sources: list[str] = []

    for workbook in workbooks:
        try:
            trial_ids, embeddings, sentiments, skipped_rows = read_workbook(workbook)
            all_ids.extend(trial_ids)
            all_embeddings.append(embeddings)
            all_sentiments.extend(sentiments)
            sources.extend([workbook.name] * len(trial_ids))
            print(
                f"Loaded {len(trial_ids):,} trials from {workbook.name}; "
                f"skipped {skipped_rows:,} malformed rows"
            )
        except ValueError as err:
            print(f"Skipping problematic file {workbook.name}: {err}")

    if not all_embeddings:
        raise ValueError("No valid trial data could be loaded from any workbook.")

    feature_matrix = np.vstack(all_embeddings)
    labels = np.asarray(all_sentiments, dtype=str)
    label_names, label_ids = np.unique(labels, return_inverse=True)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    prepared_payload = {
        "nct_id": np.asarray(all_ids, dtype=str),
        "X": feature_matrix,
        "y": label_ids.astype(np.int64),
        "label_names": label_names,
        "source_workbook": np.asarray(sources, dtype=str),
    }

    is_valid, validation_msg = validate_npz_schema(prepared_payload)
    if not is_valid:
        print(f"WARNING: Output validation failed ({validation_msg}). Proceeding with safe fallback structure.")

    np.savez_compressed(
        output_path,
        nct_id=np.asarray(all_ids, dtype=str),
        X=feature_matrix,
        y=label_ids.astype(np.int64),
        label_names=label_names,
        source_workbook=np.asarray(sources, dtype=str),
    )
    print(
        f"Saved {len(all_ids):,} trials with {feature_matrix.shape[1]} features to {output_path}"
    )


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Import clinical-trial embedding workbooks into an NPZ dataset."
    )
    parser.add_argument("input_directory", type=Path)
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("clinical/data/trials.npz"),
    )
    arguments = parser.parse_args()
    main(arguments.input_directory, arguments.output)
